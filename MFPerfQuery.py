#! /usr/bin/env python

import pprint
import requests
import re
import bs4
import openpyxl
import validators

from pprint import pprint
from datetime import date

#Function to write to a cell in Excel
def write_excel(f_file,f_sheetname,f_rownum,f_colnum,f_data):
    f_workbook = openpyxl.load_workbook(f_file)
    f_sheet = f_workbook[f_sheetname]
    f_sheet.cell(row=f_rownum, column=f_colnum).value = f_data
    f_workbook.save(f_file)

#Function to find Category
def find_cat(f_list):
    f_table = f_list.find_all("span","sub_category_text")
    f_cat = f_table[0].text
    return f_cat

#Function to find Total Expense Ratio and Category Avg
def find_ter_avg(f_list):
    f_table = f_list.find_all("div","top_section")
    f_rows = f_table[1].find_all("span","amt")
    f_ter = f_rows[1].contents[0]

    f_rows_avg = f_table[1].find_all("div","grayvalue")
    f_ter_avg = f_rows_avg[1].contents[0][2:]
    f_ter_avg = f_ter_avg.split(' ')[0]

    f_ter_value = f_ter + " vs " + f_ter_avg
    return f_ter_value

#Function to find 3Yr, 5Yr and since inception Gain and Rank in category
def find_gain_rank(f_list):
    f_y3_gain = " "
    f_y3_rank = " "
    f_y5_gain = " "
    f_y5_rank = " "
    f_incep_gain = " "
    f_incep_rank = " "
    f_table = f_list.find_all("div","data_container returns_table table-responsive")
    f_rows = f_table[0].find_all("td")
    for counter in f_rows:
        if(re.search('5 Year',repr(counter))):
            f_y5_gain = f_rows[f_rows.index(counter)+4].contents[0]
            f_y5_rank = f_rows[f_rows.index(counter)+6].contents[0]
        if(re.search('3 Year',repr(counter))):
            f_y3_gain = f_rows[f_rows.index(counter)+4].contents[0]
            f_y3_rank = f_rows[f_rows.index(counter)+6].contents[0]
        if(re.search('Since Inception',repr(counter))):
            f_incep_gain = f_rows[f_rows.index(counter)+4].contents[0]
            f_incep_rank = f_rows[f_rows.index(counter)+6].contents[0]
    
    f_gain_rank = [f_y3_gain,f_y3_rank,f_y5_gain,f_y5_rank,f_incep_gain,f_incep_rank]
    return f_gain_rank

#Function to find CRISIL Rank
def find_crisil_rank(f_list):
    f_table = f_list.find_all("div",re.compile('muttxtdn*'))
    if len(f_table) == 0:
        f_crisil = "No"
    else:
        f_attr = f_table[0].attrs
        f_crisil = f_attr['class'][1]
    return f_crisil

#Function to find Risk Rating
def find_risk_rating(f_list):
    f_table = f_list.find_all("div","meter_graph")
    f_rows = f_table[0].find_all("span","status")
    f_risk = f_rows[0].contents[0]
    return f_risk

#Function to validate URLs and responses for malformed URLs, non-200 status codes and unexpected URLs
def check_url(f_url):
    if validators.url(f_url):
        f_result = session_requests.get(f_url)
        f_code = f_result.status_code
        if f_code == 200:
            f_soup = bs4.BeautifulSoup(f_result.content,"html.parser")
            f_tag = f_soup.find("ul","bred_list")
            try:
                f_tag2 = f_tag.contents[len(f_tag.contents)-2].contents[0]
            except AttributeError:
                f_tag2 = " "
            if "Plan" in f_tag2:
                f_message = "Success"
                return (f_message,f_soup)
            else:
                f_message = "Random URL"
                return (f_message," ")
        else:
            f_message = "Incorrect URL. Received HTTP {} response code" .format(f_code)
            return(f_message," ")
    else:
        f_message = "Invalid URL"
        return(f_message," ")

session_requests = requests.session()
URL = []
soup = []
col_letter = []
max_rows = 0

print("""Choose your option:
1: To get performance of a single mutual fund by a URL
2: To get performance of multiple mutual funds from an Excel workbook""")
choice = input()

if choice.isdigit():
    if int(choice) == 1:
        get_url = input("Enter the fund URL from MoneyControl: ")
        message, result = check_url(get_url)
        if message == "Success":
            URL.insert(0,get_url)
            soup.insert(0,result)
        else:
            print(message)
            quit()
    elif int(choice) == 2:
        get_url = input("Enter the full path of the Excel (xlsx) workbook: ")
        try:
            the_file = openpyxl.load_workbook(get_url)
        except FileNotFoundError:
            print("Workbook not found")
            quit()
        sheet_name = the_file.sheetnames[0] #Load first sheet for reading URLs and writing results
        max_rows = the_file[sheet_name].max_row
        max_cols = the_file[sheet_name].max_column
        if sheet_name == "Mutual Funds": #Ensure this is the sheet we want
        #Find Column number for TER, CRISIL, Risk, Gains, Ranks and Category
            for i in range(1,max_cols+1):
                cell_value = the_file[sheet_name].cell(row=1, column=i).value
                if cell_value == "TER % vs Avg":
                    col_letter.insert(0,i)
                if cell_value == "3 YR Return":
                    col_letter.insert(1,i)
                if cell_value == "3 YR Rank / Cat Funds":
                    col_letter.insert(2,i)
                if cell_value == "5 YR Return":
                    col_letter.insert(3,i)
                if cell_value == "5 YR Rank / Cat Funds":
                    col_letter.insert(4,i)
                if cell_value == "Since Inception Return":
                    col_letter.insert(5,i)
                if cell_value == "Since Inception Rank / Cat Funds":
                    col_letter.insert(6,i)
                if cell_value == "Updated On":
                    col_letter.insert(7,i)
                if cell_value == "CRISIL Star Rank":
                    col_letter.insert(8,i)
                if cell_value == "Risk Rating":
                    col_letter.insert(9,i)
                if cell_value == "Category":
                    col_letter.insert(10,i)
        else:
            print("{} sheet not found" .format(sheet_name))
            quit()
    else:
        print("Choice not found")
        quit()
else:
    print("Invalid choice")
    quit()

#Loop to iterate through each row and get URL from column 1, response and parse it into soup
for rows in range(1,max_rows+1):
    if(the_file[sheet_name].cell(row=rows,column=1).value):
        try:
            get_url = the_file[sheet_name].cell(row=rows, column=1).hyperlink.target
            message, result = check_url(get_url)
            if message == "Success":
                URL.insert(rows-1,get_url)
                soup.insert(rows-1,result)
            else:
                URL.insert(rows-1,message)
                soup.insert(rows-1, " ")
        except AttributeError:
            URL.insert(rows-1, "No URL in cell")
            soup.insert(rows-1, " ")
    else:
        URL.insert(rows-1, "Blank cell")
        soup.insert(rows-1, " ")

#Iterate through the response content list in soup to get following values
#TER for fund vs category average, 3 year, 5 year and since inception performance for gain and rank in category
cat = []
ter_value = []
gain_rank = []
y3_gain = []
y3_rank = []
y5_gain = []
y5_rank = []
incep_gain = []
incep_rank = []
crisil = []
risk = []
today = date.today()

for lines in range(1,len(soup)+1):
    #Block to call functions to get the required fund performance parameters
    if soup[lines-1] != " ":
        ter_value.insert(lines-1,find_ter_avg(soup[lines-1]))
        gain_rank = find_gain_rank(soup[lines-1])
        y3_gain.insert(lines-1,gain_rank[0])
        y3_rank.insert(lines-1,gain_rank[1])
        y5_gain.insert(lines-1,gain_rank[2])
        y5_rank.insert(lines-1,gain_rank[3])
        incep_gain.insert(lines-1,gain_rank[4])
        incep_rank.insert(lines-1,gain_rank[5])
        crisil.insert(lines-1,find_crisil_rank(soup[lines-1]))
        risk.insert(lines-1,find_risk_rating(soup[lines-1]))
        cat.insert(lines-1,find_cat(soup[lines-1]))
    else:
        ter_value.insert(lines-1, " ")
        y3_gain.insert(lines-1," ")
        y3_rank.insert(lines-1," ")
        y5_gain.insert(lines-1," ")
        y5_rank.insert(lines-1," ")
        incep_gain.insert(lines-1," ")
        incep_rank.insert(lines-1," ")
        crisil.insert(lines-1," ")
        risk.insert(lines-1," ")
        cat.insert(lines-1, " ")
    #Block to write values to Excel or print them depending on choice
    if int(choice) == 2:
        #Skip writing when data is not available and skip row 1 since it has the Column headers
        if lines != 1 and soup[lines-1] != " ":
            if col_letter[0] != " ":
                write_excel(get_url,sheet_name,lines,col_letter[0],ter_value[lines-1])
            if col_letter[1] != " ":
                write_excel(get_url,sheet_name,lines,col_letter[1],y3_gain[lines-1])
            if col_letter[2] != " ":
                write_excel(get_url,sheet_name,lines,col_letter[2],y3_rank[lines-1])
            if col_letter[3] != " ":
                write_excel(get_url,sheet_name,lines,col_letter[3],y5_gain[lines-1])
            if col_letter[4] != " ":
                write_excel(get_url,sheet_name,lines,col_letter[4],y5_rank[lines-1])
            if col_letter[5] != " ":
                write_excel(get_url,sheet_name,lines,col_letter[5],incep_gain[lines-1])
            if col_letter[6] != " ":
                write_excel(get_url,sheet_name,lines,col_letter[6],incep_rank[lines-1])
            if col_letter[7] != " ":
                write_excel(get_url,sheet_name,lines,col_letter[7],today.strftime("%d/%m/%Y"))
            if col_letter[8] != " ":
                write_excel(get_url,sheet_name,lines,col_letter[8],crisil[lines-1])
            if col_letter[9] != " ":
                write_excel(get_url,sheet_name,lines,col_letter[9],risk[lines-1])
            if col_letter[10] != " ":
                write_excel(get_url,sheet_name,lines,col_letter[10],cat[lines-1])
    else:
        #We know choice = 1 i.e. URL based query so print out details
        print("Category                      = {}" .format(cat[lines-1]))
        print("CRISIL Star Rating            = {} star" .format(crisil[lines-1]))
        print("Risk Rating                   = {}" .format(risk[lines-1]))
        print("TER % vs AVG                  = {}" .format(ter_value[lines-1]))
        print("3 Year Gain and Rank          = {} and {}" .format(y3_gain[lines-1],y3_rank[lines-1]))
        print("5 Year Gain and Rank          = {} and {}" .format(y5_gain[lines-1],y5_rank[lines-1]))
        print("Since Inception Gain and Rank = {} and {}" .format(incep_gain[lines-1],incep_rank[lines-1]))
